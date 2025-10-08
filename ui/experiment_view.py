import customtkinter as ctk
from pathlib import Path
from tkinter import filedialog
from openpyxl import load_workbook
import csv
# pandas optional: not required for current readers (openpyxl/csv used)
pd = None


class ExperimentSection(ctk.CTkFrame):
    def __init__(self, master, on_change=None, on_send=None):
        super().__init__(master, corner_radius=12)
        self.on_change = on_change
        self.on_send = on_send
        self._selected_files: dict[str, set[str]] = {}
        self._metrics_settings: dict = {
            "header": True,
            "has_time": False,
            "time_col": "",
            "selected_cols": set(),
        }
        self._config_settings: dict = {
            "flatten": False,
        }
        self._raw_data_settings: dict = {
            "send_minio": True,
            "save_locally": False,
            "local_path": "",
        }
        # CSV separators per selector (persisted)
        self._csv_separators: dict[str, str] = {
            "config": ",",
            "metrics": ",",
            "results": ",",
        }
        # batch sending controls (not persisted)
        self._batch_enable = False
        self._batch_selected: set[str] = set()
        self._allowed_tabular_suffixes = (".json", ".csv", ".xlsx", ".xlsm")
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=0)
        self.grid_rowconfigure(999, weight=1)

        ctk.CTkLabel(self, text="Experiment Files", font=("Segoe UI", 18, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(12, 8)
        )

        ctk.CTkLabel(self, text="Experiment folder").grid(row=1, column=0, sticky="w", padx=12)
        self.folder_entry = ctk.CTkEntry(self, placeholder_text="Select a folder…")
        self.folder_entry.grid(row=1, column=1, sticky="ew", padx=6, pady=6)
        ctk.CTkButton(self, text="Browse…", command=self.choose_folder).grid(
            row=1, column=2, sticky="e", padx=(6, 12), pady=6
        )

        self.file_menus: dict[str, ctk.CTkOptionMenu] = {}
        self.sheet_menus: dict[str, ctk.CTkOptionMenu] = {}
        self._keys = [
            ("config", "Config"),
            ("results", "Results"),
            ("metrics", "Metrics"),
            ("raw_data", "Raw data"),
            ("artifacts", "Artifacts"),
        ]
        for idx, (key, label) in enumerate(self._keys, start=2):
            ctk.CTkLabel(self, text=label).grid(row=idx, column=0, sticky="w", padx=12)
            # For config: no "None" choice; others keep "None"
            initial_values = [""] if key == "config" else ["None"]
            file_menu = ctk.CTkOptionMenu(
                self, values=initial_values, dynamic_resizing=False, width=320,
                command=lambda v, k=key: self.on_file_changed(k, v)
            )
            file_menu.grid(row=idx, column=1, sticky="ew", padx=6, pady=6)
            file_menu.set("" if key == "config" else "None")
            self.file_menus[key] = file_menu

            sheet_menu = ctk.CTkOptionMenu(self, values=[""], dynamic_resizing=False, width=200, command=lambda v, k=key: self.on_sheet_changed(k, v))
            sheet_menu.grid(row=idx, column=2, sticky="ew", padx=(6, 12), pady=6)
            sheet_menu.grid_remove()
            self.sheet_menus[key] = sheet_menu

        # container for dynamic per-selector sections (two-column layout, small margins)
        self.details_container = ctk.CTkFrame(self, corner_radius=12)
        base_row = 2 + len(self._keys) + 1
        self.details_container.grid(row=base_row, column=0, columnspan=3, sticky="nsew", padx=6, pady=(6, 6))
        self.details_container.grid_columnconfigure(0, weight=1)
        self.details_container.grid_columnconfigure(1, weight=1)

        # Batch toggle and siblings list
        batch_row = base_row + 1
        self.batch_enable_var = ctk.BooleanVar(value=False)
        def on_batch_toggle():
            self._batch_enable = bool(self.batch_enable_var.get())
            self._render_batch_checkboxes()
            if callable(self.on_change):
                self.on_change()
        ctk.CTkCheckBox(self, text="Send multiple experiments", variable=self.batch_enable_var, command=on_batch_toggle).grid(
            row=batch_row, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 4)
        )
        self.batch_container = ctk.CTkFrame(self, corner_radius=8)
        self.batch_container.grid(row=batch_row + 1, column=0, columnspan=3, sticky="nsew", padx=6, pady=(0, 6))
        # hide container by default so it doesn't reserve space
        try:
            self.batch_container.grid_remove()
        except Exception:
            pass
        self.batch_container.grid_columnconfigure(0, weight=1)
        self._render_batch_checkboxes()

        # actions row: Send experiment button inside the section, below batch
        actions_row = ctk.CTkFrame(self, fg_color="transparent")
        actions_row.grid(row=batch_row + 2, column=0, columnspan=3, sticky="ew", padx=6, pady=(0, 2))
        actions_row.grid_columnconfigure(2, weight=1)
        send_btn = ctk.CTkButton(actions_row, text="Send experiment", width=180, height=36, command=self._on_send_click)
        send_btn.grid(row=0, column=2, sticky="e", padx=(0, 6), pady=(2, 2))

        # status labels: one for file/cards errors, one for send result
        self.status = ctk.CTkLabel(self, text="", wraplength=520, justify="left")
        self.status.grid(row=batch_row + 3, column=0, columnspan=3, sticky="ew", padx=12, pady=(2, 6))
        self.send_status = ctk.CTkLabel(self, text="", wraplength=520, justify="left")
        self.send_status.grid(row=batch_row + 4, column=0, columnspan=3, sticky="ew", padx=12, pady=(2, 4))

    def _on_send_click(self):
        try:
            if callable(self.on_send):
                self.on_send()
        except Exception:
            pass

    # --- IO ---
    def get_prefs(self) -> dict:
        data = {"experiment_folder": self.folder_entry.get().strip()}
        for key, _ in self._keys:
            data[f"{key}_name"] = (self.file_menus[key].get() or "").strip()
            data[f"{key}_sheet"] = (self.sheet_menus[key].get() or "").strip()
            if key in ("raw_data", "artifacts"):
                selected = sorted(list(self._selected_files.get(key, set())))
                data[f"{key}_files"] = selected
        # metrics settings persistence
        data["metrics_header"] = int(bool(self._metrics_settings.get("header", True)))
        data["metrics_has_time"] = int(bool(self._metrics_settings.get("has_time", False)))
        data["metrics_time_col"] = self._metrics_settings.get("time_col", "")
        data["metrics_selected_cols"] = sorted(list(self._metrics_settings.get("selected_cols", set())))
        # config settings persistence
        data["config_flatten"] = int(bool(self._config_settings.get("flatten", False)))
        # raw_data settings persistence
        data["raw_data_send_minio"] = int(bool(self._raw_data_settings.get("send_minio", True)))
        data["raw_data_save_locally"] = int(bool(self._raw_data_settings.get("save_locally", False)))
        data["raw_data_local_path"] = self._raw_data_settings.get("local_path", "")
        # CSV separators
        data["config_sep"] = self._csv_separators.get("config", ",")
        data["metrics_sep"] = self._csv_separators.get("metrics", ",")
        data["results_sep"] = self._csv_separators.get("results", ",")
        # compute list of folders per batch toggle
        folders_list: list[str] = []
        base_folder = (self.folder_entry.get() or "").strip()
        try:
            if self._batch_enable:
                # list siblings of selected folder
                if base_folder:
                    p = Path(base_folder)
                    parent = p.parent if p.exists() else None
                    if parent and parent.exists():
                        # if no selection yet, default to all siblings
                        if not self._batch_selected:
                            try:
                                for d in parent.iterdir():
                                    if d.is_dir() and not d.name.startswith("."):
                                        self._batch_selected.add(d.name)
                            except Exception:
                                pass
                        for name in sorted(list(self._batch_selected)):
                            folders_list.append(str((parent / name).resolve()))
            else:
                if base_folder:
                    folders_list = [str(Path(base_folder).resolve())]
        except Exception:
            folders_list = [base_folder] if base_folder else []
        data["experiment_folders"] = folders_list
        return data

    def set_prefs(self, data: dict):
        exp_dir = data.get("experiment_folder", "")
        if exp_dir:
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, exp_dir)
        # populate menu values
        self.refresh_items(preserve_selection=False)
        # restore selection
        for key, _ in self._keys:
            name = data.get(f"{key}_name", "") or "None"
            try:
                self.file_menus[key].set(name)
            except Exception:
                self.file_menus[key].set(name)
            self.update_sheet_menu_for(key)
        # restore sheet
        for key, _ in self._keys:
            sheet = data.get(f"{key}_sheet", "")
            if sheet:
                try:
                    self.sheet_menus[key].set(sheet)
                except Exception:
                    self.sheet_menus[key].set(sheet)
            self.update_sheet_menu_for(key)
        # restore per-folder file selections
        for key in ("raw_data", "artifacts"):
            saved = data.get(f"{key}_files", [])
            if isinstance(saved, list):
                self._selected_files[key] = set(saved)
        # restore metrics settings
        self._metrics_settings["header"] = bool(data.get("metrics_header", 1))
        self._metrics_settings["has_time"] = bool(data.get("metrics_has_time", 0))
        self._metrics_settings["time_col"] = data.get("metrics_time_col", "") or ""
        sel = data.get("metrics_selected_cols", [])
        self._metrics_settings["selected_cols"] = set(sel) if isinstance(sel, list) else set()
        # restore config settings
        self._config_settings["flatten"] = bool(data.get("config_flatten", 0))
        # restore raw_data settings
        self._raw_data_settings["send_minio"] = bool(data.get("raw_data_send_minio", 1))
        self._raw_data_settings["save_locally"] = bool(data.get("raw_data_save_locally", 0))
        self._raw_data_settings["local_path"] = data.get("raw_data_local_path", "") or ""
        # restore CSV separators
        self._csv_separators["config"] = data.get("config_sep", ",") or ","
        self._csv_separators["metrics"] = data.get("metrics_sep", ",") or ","
        self._csv_separators["results"] = data.get("results_sep", ",") or ","
        # restore experiment name
        # self.exp_name_entry.delete(0, "end")
        # self.exp_name_entry.insert(0, data.get("experiment_name", ""))
        # ensure details reflect restored prefs
        self.render_details_sections()

    def _render_batch_checkboxes(self):
        # clear
        for child in list(self.batch_container.winfo_children() if hasattr(self, 'batch_container') else []):
            try:
                child.destroy()
            except Exception:
                pass
        if not getattr(self, 'batch_container', None):
            return
        # Hide the container when disabled; show when enabled
        if not bool(self.batch_enable_var.get()):
            try:
                self.batch_container.grid_remove()
            except Exception:
                pass
            return
        else:
            try:
                self.batch_container.grid()
            except Exception:
                pass
        # build siblings list
        base_folder = (self.folder_entry.get() or "").strip()
        siblings: list[str] = []
        try:
            if base_folder:
                p = Path(base_folder)
                parent = p.parent if p.exists() else None
                if parent and parent.exists():
                    siblings = [d.name for d in parent.iterdir() if d.is_dir() and not d.name.startswith('.')]
                    siblings.sort(key=lambda n: n.lower())
        except Exception:
            siblings = []
        # default select all if nothing yet
        if not self._batch_selected:
            self._batch_selected = set(siblings)
        # render
        for i, name in enumerate(siblings):
            var = ctk.BooleanVar(value=(name in self._batch_selected))
            def _toggle(n=name, v=var):
                if v.get():
                    self._batch_selected.add(n)
                else:
                    try:
                        self._batch_selected.remove(n)
                    except KeyError:
                        pass
                if callable(self.on_change):
                    self.on_change()
            cb = ctk.CTkCheckBox(self.batch_container, text=name, variable=var, command=_toggle)
            cb.grid(row=i, column=0, sticky="w", padx=8, pady=2)

    # --- Events ---
    def choose_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, folder)
            self.refresh_items(preserve_selection=True)
        if callable(self.on_change):
            self.on_change()

    def on_file_changed(self, key: str, value: str):
        # Reset dependent selections when source file/folder changes
        try:
            if key == "metrics":
                # clear selected columns and time column; will be recomputed on render
                self._metrics_settings["selected_cols"] = set()
                self._metrics_settings["time_col"] = ""
            elif key in ("raw_data", "artifacts"):
                # clear previously selected files so defaults (all files) apply
                self._selected_files[key] = set()
        except Exception:
            pass
        self.update_sheet_menu_for(key)
        self.render_details_sections()
        if callable(self.on_change):
            self.on_change()

    def on_sheet_changed(self, key: str, value: str):
        # When a sheet is selected, simply re-render the cards
        self.render_details_sections()
        if callable(self.on_change):
            self.on_change()

    def get_full_path_for_key(self, key: str) -> Path | None:
        base_folder = self.folder_entry.get().strip()
        name = self.file_menus[key].get().strip()
        if not base_folder or not name or name == "None":
            return None
        return Path(base_folder) / name

    def update_sheet_menu_for(self, key: str):
        sheet_menu = self.sheet_menus[key]
        path = self.get_full_path_for_key(key)
        # hide if not supported tabular file
        # special-case: for raw_data and artifacts never show sheet selector
        if not path or path.is_dir() or key in ("raw_data", "artifacts") or path.suffix.lower() not in (".xlsx", ".xlsm"):
            try:
                sheet_menu.grid_remove()
            except Exception:
                pass
            sheet_menu.configure(values=[""])
            sheet_menu.set("")
            return
        # fetch sheet names
        sheets: list[str] = []
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            sheets = list(wb.sheetnames)
            wb.close()
        except Exception as e:
            self.status.configure(text=f"Could not read sheets from {path.name}: {e}")
        if not sheets:
            sheets = [""]
        try:
            sheet_menu.configure(values=sheets)
        except Exception:
            pass
        # show and keep previous selection if still present
        try:
            sheet_menu.grid()
        except Exception:
            pass
        current = sheet_menu.get() or ""
        if current and current in sheets:
            sheet_menu.set(current)
        else:
            sheet_menu.set(sheets[0] if sheets else "")

    def refresh_items(self, preserve_selection: bool = True):
        base_folder = self.folder_entry.get().strip()
        all_items = self._list_items()
        restricted = {"config", "metrics", "results"}
        for key, _ in self._keys:
            current = (self.file_menus[key].get() or "") if preserve_selection else ""
            # Build values list per key
            if key in restricted and base_folder:
                try:
                    base = Path(base_folder)
                    filtered = [n for n in all_items if (base / n).is_file() and (base / n).suffix.lower() in self._allowed_tabular_suffixes]
                except Exception:
                    filtered = []
                # For config, remove the "None" option entirely
                values = (filtered if key == "config" else ["None"] + filtered)
            else:
                values = ["None"] + list(all_items)

            # Configure menu
            try:
                self.file_menus[key].configure(values=values)
            except Exception:
                pass

            # Reset invalid current selections for restricted keys
            if key == "config":
                target_value = current if (current and current in values) else (values[0] if values else "")
            else:
                target_value = current if (current and current in values) else "None"
            try:
                self.file_menus[key].set(target_value)
            except Exception:
                self.file_menus[key].set(target_value)

        # refresh details after items update
        self.render_details_sections()

    def _list_items(self) -> list[str]:
        base_folder = self.folder_entry.get().strip()
        if not base_folder:
            return []
        try:
            base = Path(base_folder)
            if not base.exists() or not base.is_dir():
                return []
            names = [p.name for p in base.iterdir()]
            names.sort(key=lambda n: n.lower())
            return names
        except Exception:
            return []

    # --- Dynamic details per selector ---
    def render_details_sections(self):
        # clear any previous error message on each cards update
        try:
            self.status.configure(text="")
        except Exception:
            pass
        # clear previous
        for child in list(self.details_container.winfo_children()):
            try:
                child.destroy()
            except Exception:
                pass
        cols = 2
        idx = 0
        for key, label in self._keys:
            name = (self.file_menus[key].get() or "").strip()
            if not name or name == "None":
                continue
            path = self.get_full_path_for_key(key)
            # create a simple frame per selection in a two-column grid
            r, c = divmod(idx, cols)
            sec = ctk.CTkFrame(self.details_container, corner_radius=12)
            sec.grid(row=r, column=c, sticky="nsew", padx=6, pady=6)
            sec.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(sec, text=f"{label}", font=("Segoe UI", 14, "bold")).grid(
                row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(8, 4)
            )
            ctk.CTkLabel(sec, text="Selected file").grid(row=1, column=0, sticky="w", padx=8, pady=4)
            ctk.CTkLabel(sec, text=name).grid(row=1, column=1, sticky="w", padx=(6, 8), pady=4)
            # sheet (if visible)
            sheet = (self.sheet_menus[key].get() or "").strip()
            if sheet and not (key in ("raw_data", "artifacts")):
                ctk.CTkLabel(sec, text="Sheet").grid(row=2, column=0, sticky="w", padx=8, pady=4)
                ctk.CTkLabel(sec, text=sheet).grid(row=2, column=1, sticky="w", padx=(6, 8), pady=4)
            # CSV separator selector for config/metrics/results
            if path and path.is_file() and path.suffix.lower() == ".csv" and key in ("config", "metrics", "results"):
                sep_row = 3 if (sheet and not (key in ("raw_data", "artifacts"))) else 2
                ctk.CTkLabel(sec, text="Separator").grid(row=sep_row, column=0, sticky="w", padx=8, pady=4)
                sep_menu = ctk.CTkOptionMenu(
                    sec,
                    values=[",", ";", "|", "\\t"],
                    dynamic_resizing=False,
                    command=lambda v, k=key: self._on_sep_changed(k, v)
                )
                current_sep = self._csv_separators.get(key, ",")
                display_val = "\\t" if current_sep == "\t" else current_sep
                sep_menu.set(display_val)
                sep_menu.grid(row=sep_row, column=1, sticky="ew", padx=(6, 8), pady=4)
            # folder checklist for raw_data / artifacts
            if key in ("raw_data", "artifacts") and path and path.is_dir():
                files = []
                try:
                    files = [p.name for p in Path(path).iterdir() if p.is_file()]
                    files.sort(key=lambda n: n.lower())
                except Exception:
                    files = []
                # initialize default selection: previously saved intersected with current files; new files selected by default
                saved = self._selected_files.get(key, set())
                if saved:
                    selected = set(f for f in files if f in saved)
                    # select also new files by default
                    for f in files:
                        if f not in saved:
                            selected.add(f)
                else:
                    selected = set(files)
                self._selected_files[key] = selected
                # render checkboxes
                chk_container = ctk.CTkFrame(sec, corner_radius=8)
                chk_container.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=6, pady=(4, 6))
                # arrange in two columns if many files
                for i in range(2):
                    chk_container.grid_columnconfigure(i, weight=1)
                for i, fname in enumerate(files):
                    col = i % 2
                    rowc = i // 2
                    var = ctk.BooleanVar(value=fname in selected)
                    def _make_cmd(k=key, name=fname, v=var):
                        return lambda: self._on_file_toggle(k, name, v.get())
                    cb = ctk.CTkCheckBox(chk_container, text=fname, variable=var, command=_make_cmd())
                    cb.grid(row=rowc, column=col, sticky="w", padx=6, pady=2)
            # raw_data controls: Send Minio / Save locally + path
            if key == "raw_data":
                has_checklist = bool(path and path.is_dir())
                next_row_local = 4 if has_checklist else 2
                # Send Minio checkbox
                send_var = ctk.BooleanVar(value=bool(self._raw_data_settings.get("send_minio", True)))
                def on_send_toggle():
                    self._raw_data_settings["send_minio"] = bool(send_var.get())
                    if callable(self.on_change):
                        self.on_change()
                send_cb = ctk.CTkCheckBox(sec, text="Send Minio", variable=send_var, command=on_send_toggle)
                send_cb.grid(row=next_row_local, column=0, sticky="w", padx=8, pady=(6, 4))
                # Save locally checkbox
                save_var = ctk.BooleanVar(value=bool(self._raw_data_settings.get("save_locally", False)))
                def on_save_toggle():
                    self._raw_data_settings["save_locally"] = bool(save_var.get())
                    try:
                        entry.configure(state=("normal" if save_var.get() else "disabled"))
                        btn.configure(state=("normal" if save_var.get() else "disabled"))
                    except Exception:
                        pass
                    if callable(self.on_change):
                        self.on_change()
                save_cb = ctk.CTkCheckBox(sec, text="Save locally", variable=save_var, command=on_save_toggle)
                save_cb.grid(row=next_row_local, column=1, sticky="w", padx=8, pady=(6, 4))
                # Path selector
                def choose_path():
                    folder = filedialog.askdirectory()
                    if folder:
                        try:
                            entry.delete(0, "end")
                            entry.insert(0, folder)
                        except Exception:
                            pass
                        self._raw_data_settings["local_path"] = folder
                        if callable(self.on_change):
                            self.on_change()
                ctk.CTkLabel(sec, text="Local path").grid(row=next_row_local + 1, column=0, sticky="w", padx=8, pady=4)
                entry = ctk.CTkEntry(sec, placeholder_text="Select a folder…")
                entry.grid(row=next_row_local + 1, column=1, sticky="ew", padx=(6, 8), pady=4)
                if self._raw_data_settings.get("local_path"):
                    try:
                        entry.delete(0, "end")
                        entry.insert(0, self._raw_data_settings.get("local_path", ""))
                    except Exception:
                        pass
                btn = ctk.CTkButton(sec, text="Browse…", width=90, command=choose_path)
                btn.grid(row=next_row_local + 2, column=1, sticky="e", padx=(6, 8), pady=(0, 6))
                try:
                    entry.configure(state=("normal" if save_var.get() else "disabled"))
                    btn.configure(state=("normal" if save_var.get() else "disabled"))
                except Exception:
                    pass
            # config controls
            # show Flatten checkbox if config file is a JSON
            if key == "config" and path and path.is_file() and path.suffix.lower() == ".json":
                # decide next available row: 3 if sheet displayed, else 2
                next_row_local = 3 if (self.sheet_menus[key].get().strip() and not (key in ("raw_data", "artifacts"))) else 2
                flatten_var = ctk.BooleanVar(value=bool(self._config_settings.get("flatten", False)))
                def on_flatten_toggle():
                    self._config_settings["flatten"] = bool(flatten_var.get())
                    if callable(self.on_change):
                        self.on_change()
                flatten_cb = ctk.CTkCheckBox(sec, text="Flatten", variable=flatten_var, command=on_flatten_toggle)
                flatten_cb.grid(row=next_row_local, column=0, sticky="w", padx=8, pady=4)
            # metrics DataFrame controls
            if key == "metrics" and path and path.is_file():
                col_names, data_rows = self._read_tabular(path, sheet)
                # defaults for selected columns: if none saved, select all except time col
                if not self._metrics_settings.get("selected_cols"):
                    self._metrics_settings["selected_cols"] = set(col_names)
                # Header checkbox
                header_var = ctk.BooleanVar(value=bool(self._metrics_settings.get("header", True)))
                def on_header_toggle():
                    self._metrics_settings["header"] = bool(header_var.get())
                    # reset selected cols to match new headers
                    cols2, _ = self._read_tabular(path, sheet)
                    self._metrics_settings["selected_cols"] = set(cols2)
                    # if time col no longer exists, reset
                    if self._metrics_settings.get("time_col") not in cols2:
                        self._metrics_settings["time_col"] = ""
                    self.render_details_sections()
                    if callable(self.on_change):
                        self.on_change()
                header_cb = ctk.CTkCheckBox(sec, text="Column header", variable=header_var, command=on_header_toggle)
                header_cb.grid(row=3, column=0, sticky="w", padx=8, pady=(6, 4))

                # Time column checkbox
                has_time_var = ctk.BooleanVar(value=bool(self._metrics_settings.get("has_time", False)))
                def on_has_time_toggle():
                    self._metrics_settings["has_time"] = bool(has_time_var.get())
                    if not has_time_var.get():
                        self._metrics_settings["time_col"] = ""
                    self.render_details_sections()
                    if callable(self.on_change):
                        self.on_change()
                time_cb = ctk.CTkCheckBox(sec, text="Time column", variable=has_time_var, command=on_has_time_toggle)
                time_cb.grid(row=3, column=1, sticky="w", padx=8, pady=(6, 4))

                next_row = 4
                current_cols = list(col_names)
                # Time column selector if enabled
                if has_time_var.get():
                    time_values = current_cols
                    time_menu = ctk.CTkOptionMenu(sec, values=time_values, dynamic_resizing=False,
                                                  command=lambda v: self._on_metrics_time_column_changed(v))
                    # set current
                    if self._metrics_settings.get("time_col") in time_values:
                        time_menu.set(self._metrics_settings.get("time_col"))
                    elif time_values:
                        time_menu.set(time_values[0])
                        self._metrics_settings["time_col"] = time_values[0]
                    ctk.CTkLabel(sec, text="Time column").grid(row=next_row, column=0, sticky="w", padx=8, pady=4)
                    time_menu.grid(row=next_row, column=1, sticky="ew", padx=(6, 8), pady=4)
                    next_row += 1

                # Columns checklist (exclude time column if set)
                cols_to_list = [c for c in current_cols if c != self._metrics_settings.get("time_col", "")]
                chk_container = ctk.CTkFrame(sec, corner_radius=8)
                chk_container.grid(row=next_row, column=0, columnspan=2, sticky="nsew", padx=6, pady=(4, 6))
                for i in range(2):
                    chk_container.grid_columnconfigure(i, weight=1)
                for i, cname in enumerate(cols_to_list):
                    col = i % 2
                    rowc = i // 2
                    var = ctk.BooleanVar(value=(cname in self._metrics_settings.get("selected_cols", set())))
                    def _make_cmd2(name=cname, v=var):
                        return lambda: self._on_metrics_column_toggle(name, v.get())
                    cb = ctk.CTkCheckBox(chk_container, text=cname, variable=var, command=_make_cmd2())
                    cb.grid(row=rowc, column=col, sticky="w", padx=6, pady=2)
            idx += 1

    def _on_file_toggle(self, key: str, filename: str, is_selected: bool):
        sel = self._selected_files.get(key, set())
        if is_selected:
            sel.add(filename)
        else:
            try:
                sel.remove(filename)
            except KeyError:
                pass
        self._selected_files[key] = sel
        if callable(self.on_change):
            self.on_change()

    # --- Metrics helpers ---
    def _read_tabular(self, path: Path, sheet: str) -> tuple[list[str], list[list[object]]]:
        cols: list[str] = []
        rows: list[list[object]] = []
        try:
            if path.suffix.lower() in (".xlsx", ".xlsm"):
                wb = load_workbook(filename=str(path), read_only=True, data_only=True)
                ws = None
                if sheet and sheet in wb.sheetnames:
                    ws = wb[sheet]
                else:
                    ws = wb[wb.sheetnames[0]]
                data_iter = ws.iter_rows(values_only=True)
                for i, row in enumerate(data_iter):
                    if i == 0 and bool(self._metrics_settings.get("header", True)):
                        cols = [str(c) if c is not None else f"col{idx}" for idx, c in enumerate(list(row))]
                    else:
                        rows.append(list(row))
                wb.close()
                # if no header, generate from max row length
                if not cols:
                    max_len = max((len(r) for r in rows), default=0)
                    cols = [str(i) for i in range(max_len)]
            elif path.suffix.lower() == ".csv":
                # Use selected separator for metrics preview
                sep = self._csv_separators.get("metrics", ",")
                with open(path, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f, delimiter=("\t" if sep == "\t" else sep))
                    for i, row in enumerate(reader):
                        if i == 0 and bool(self._metrics_settings.get("header", True)):
                            cols = [str(c) for c in row]
                        else:
                            rows.append(row)
                if not cols:
                    max_len = max((len(r) for r in rows), default=0)
                    cols = [str(i) for i in range(max_len)]
            else:
                # unsupported -> empty
                cols, rows = [], []
        except Exception as e:
            try:
                self.status.configure(text=f"Error reading metrics: {e}")
            except Exception:
                pass
        return cols, rows

    def _on_sep_changed(self, key: str, display_value: str):
        sep = "\t" if display_value == "\\t" else display_value
        self._csv_separators[key] = sep
        # re-render metrics preview immediately to reflect new parsing
        if key == "metrics":
            self.render_details_sections()
        if callable(self.on_change):
            self.on_change()

    def _on_metrics_time_column_changed(self, col_name: str):
        self._metrics_settings["time_col"] = col_name or ""
        if col_name:
            # ensure selected cols include all except time col if none set
            if not self._metrics_settings.get("selected_cols"):
                self._metrics_settings["selected_cols"] = set()
        self.render_details_sections()
        if callable(self.on_change):
            self.on_change()

    def _on_metrics_column_toggle(self, col_name: str, is_selected: bool):
        sel = self._metrics_settings.get("selected_cols", set())
        if is_selected:
            sel.add(col_name)
        else:
            try:
                sel.remove(col_name)
            except KeyError:
                pass
        self._metrics_settings["selected_cols"] = sel
        if callable(self.on_change):
            self.on_change()

